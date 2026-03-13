from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Annotated, Any
from datetime import datetime, timedelta, timezone
import base64
import hashlib
import json
import logging
import os
import secrets
import sqlite3
import urllib.error
import urllib.request
from io import BytesIO

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("appcot")

from fastapi import Cookie, FastAPI, HTTPException, Query, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, ConfigDict, Field
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from starlette.background import BackgroundTask
import uvicorn


BASE_DIR = Path(__file__).resolve().parent


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if not key:
            continue
        if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
            value = value[1:-1]
        os.environ.setdefault(key, value)


load_env_file(BASE_DIR / ".env")

DB_PATH = BASE_DIR / "app.db"
PRICES_PATH = BASE_DIR / "materiales" / "prices.json"
TARGET_SHEET_NAME = "Formato No. 1"
MAX_ITEMS = 4
DEFAULT_COMMISSION_FACTOR = 1.15
SESSION_COOKIE_NAME = "appcot_session"
RESEND_API_URL = "https://api.resend.com/emails"
RESEND_API_KEY = os.getenv("RESEND_API_KEY", "").strip()
EMAIL_FROM = os.getenv("EMAIL_FROM", "onboarding@resend.dev").strip()
EMAIL_REPLY_TO = os.getenv("EMAIL_REPLY_TO", "").strip()
EMAIL_MODE = os.getenv("EMAIL_MODE", "resend" if RESEND_API_KEY else "stub").strip().lower()

SPANISH_MONTHS = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}


def utcnow() -> datetime:
    return datetime.now(timezone.utc)


def iso_now() -> str:
    return utcnow().isoformat()


def hash_token(raw: str) -> str:
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def db_connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db() -> None:
    logger.info("Initializing database at %s", DB_PATH)
    conn = db_connect()
    try:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              email TEXT NOT NULL UNIQUE,
              name TEXT,
              role TEXT NOT NULL DEFAULT 'cotizador',
              is_active INTEGER NOT NULL DEFAULT 1,
              created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS magic_link_tokens (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              user_id INTEGER NOT NULL,
              token_hash TEXT NOT NULL UNIQUE,
              expires_at TEXT NOT NULL,
              used_at TEXT,
              created_at TEXT NOT NULL,
              FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS session_tokens (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              user_id INTEGER NOT NULL,
              token_hash TEXT NOT NULL UNIQUE,
              expires_at TEXT NOT NULL,
              created_at TEXT NOT NULL,
              FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS cotizaciones (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              status TEXT NOT NULL DEFAULT 'pending',
              full_name TEXT NOT NULL,
              company_name TEXT NOT NULL,
              emails_json TEXT NOT NULL,
              line_product TEXT NOT NULL,
              monthly_meters REAL,
              product_name TEXT NOT NULL DEFAULT 'Flex GL',
              commission_factor REAL NOT NULL DEFAULT 1.15,
              review_notes TEXT,
              created_at TEXT NOT NULL,
              updated_at TEXT NOT NULL,
              approved_at TEXT,
              approved_by_user_id INTEGER,
              FOREIGN KEY(approved_by_user_id) REFERENCES users(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS cotizacion_items (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              cotizacion_id INTEGER NOT NULL,
              position INTEGER NOT NULL,
              type TEXT NOT NULL,
              calibre TEXT NOT NULL,
              width REAL NOT NULL,
              barrier_type TEXT NOT NULL,
              seal_type TEXT NOT NULL,
              price_override_p100 REAL,
              FOREIGN KEY(cotizacion_id) REFERENCES cotizaciones(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS email_stub_logs (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              cotizacion_id INTEGER NOT NULL,
              recipients_json TEXT NOT NULL,
              subject TEXT NOT NULL,
              body_preview TEXT NOT NULL,
              created_at TEXT NOT NULL,
              FOREIGN KEY(cotizacion_id) REFERENCES cotizaciones(id) ON DELETE CASCADE
            );
            """
        )
        conn.commit()
        logger.info("Database initialized OK")
    finally:
        conn.close()


def remove_temp_file(path: str) -> None:
    try:
        Path(path).unlink(missing_ok=True)
    except OSError:
        pass


def get_today_date_spanish() -> str:
    today = datetime.now()
    month_name = SPANISH_MONTHS[today.month]
    return f"{today.day} de {month_name} de {today.year}"


def parse_excel_number(value: str | int | float | None) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        parsed = float(text)
        return int(parsed) if parsed.is_integer() else parsed
    except ValueError:
        return None


def clear_product_row(sheet, row: int, start_col: str = "B", end_col: str = "L") -> None:
    for col_ord in range(ord(start_col), ord(end_col) + 1):
        sheet[f"{chr(col_ord)}{row}"] = None


def get_material_record(material_name: str) -> dict | None:
    if not PRICES_PATH.exists():
        return None
    try:
        with PRICES_PATH.open("r", encoding="utf-8") as file:
            prices_data = json.load(file)
    except Exception:  # noqa: BLE001
        return None

    materiales = prices_data.get("materiales")
    if isinstance(materiales, list):
        return next(
            (item for item in materiales if str(item.get("name", "")).strip().lower() == material_name.strip().lower()),
            None,
        )

    # Backward compatibility with previous shape
    if isinstance(materiales, dict):
        tapas = materiales.get("tapas", [])
        return next(
            (item for item in tapas if str(item.get("name", "")).strip().lower() == material_name.strip().lower()),
            None,
        )
    return None


def get_milesimas_for_material(material_name: str, calibre_micras: str) -> str | None:
    material = get_material_record(material_name)
    if material is None:
        return None
    try:
        micras_key = str(int(float(calibre_micras.strip())))
    except ValueError:
        return None
    values = material.get("prices_by_micras", {}).get(micras_key)
    if values is None or values.get("espesor_milesimas") is None:
        return None
    return f"{float(values['espesor_milesimas']):.1f}"


def get_price_for_material(material_name: str, calibre_micras: str) -> float | None:
    material = get_material_record(material_name)
    if material is None:
        return None
    try:
        micras_key = str(int(float(calibre_micras.strip())))
    except ValueError:
        return None
    values = material.get("prices_by_micras", {}).get(micras_key)
    if values is None or values.get("price") is None:
        return None
    try:
        return float(values["price"])
    except (TypeError, ValueError):
        return None


def send_email_with_resend(
    recipients: list[str],
    subject: str,
    html_body: str,
    text_body: str | None = None,
    attachments: list[dict[str, str]] | None = None,
) -> dict[str, Any]:
    if not RESEND_API_KEY:
        raise RuntimeError("RESEND_API_KEY is not configured")
    if not recipients:
        raise RuntimeError("No recipients provided")

    payload: dict[str, Any] = {
        "from": EMAIL_FROM,
        "to": recipients,
        "subject": subject,
        "html": html_body,
    }
    if text_body:
        payload["text"] = text_body
    if EMAIL_REPLY_TO:
        payload["reply_to"] = EMAIL_REPLY_TO
    if attachments:
        payload["attachments"] = attachments

    body = json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(
        RESEND_API_URL,
        data=body,
        headers={
            "Authorization": f"Bearer {RESEND_API_KEY}",
            "Content-Type": "application/json",
            "User-Agent": "appcot-backend/0.1",
        },
        method="POST",
    )
    logger.info("Sending email via Resend to %s | subject: %s", recipients, subject)
    try:
        with urllib.request.urlopen(request, timeout=20) as response:  # noqa: S310
            response_body = response.read().decode("utf-8")
            parsed = json.loads(response_body) if response_body else {}
            provider_id = parsed.get("id")
            logger.info("Resend accepted email | provider_id=%s | to=%s", provider_id, recipients)
            return {
                "status_code": response.status,
                "provider_id": provider_id,
                "raw_response": parsed,
            }
    except urllib.error.HTTPError as exc:
        details = exc.read().decode("utf-8", errors="replace")
        logger.error("Resend HTTP error %s | to=%s | body=%s", exc.code, recipients, details)
        raise RuntimeError(f"Resend HTTP {exc.code}: {details}") from exc
    except urllib.error.URLError as exc:
        logger.error("Resend connection error | to=%s | reason=%s", recipients, exc.reason)
        raise RuntimeError(f"Resend connection error: {exc.reason}") from exc


class QuoteItemPayload(BaseModel):
    type: str
    calibre: str
    width: float
    barrierType: str = "alta"
    sealType: str = "hermetico"
    priceOverrideP100: float | None = None


class CotizacionCreateRequest(BaseModel):
    model_config = ConfigDict(populate_by_name=True)
    full_name: str = Field(alias="fullName")
    company_name: str = Field(alias="companyName")
    emails: list[str] = Field(default_factory=list)
    line_product: str = Field(alias="lineProduct")
    monthly_meters: float | None = Field(default=None, alias="monthlyMeters")
    product_name: str = Field(default="Flex GL", alias="productName")
    items: list[QuoteItemPayload] = Field(default_factory=list, max_length=MAX_ITEMS)


class MagicLinkRequest(BaseModel):
    email: str
    name: str | None = None


class MagicLinkVerifyRequest(BaseModel):
    token: str


class CotizacionUpdateRequest(BaseModel):
    commissionFactor: float | None = None
    reviewNotes: str | None = None
    lineProduct: str | None = None
    monthlyMeters: float | None = None
    emails: list[str] | None = None
    items: list[dict[str, Any]] | None = None


app = FastAPI(title="AppCot API", version="0.2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost", "http://localhost:3000", "http://localhost:3001", "http://localhost:3002"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

logger.info("Starting AppCot API | DB=%s | EMAIL_MODE=%s | EMAIL_FROM=%s", DB_PATH, EMAIL_MODE, EMAIL_FROM)
init_db()


def get_current_user(session_cookie: str | None = Cookie(default=None, alias=SESSION_COOKIE_NAME)) -> sqlite3.Row:
    if not session_cookie:
        raise HTTPException(status_code=401, detail="Not authenticated")
    token_hash = hash_token(session_cookie)
    conn = db_connect()
    try:
        row = conn.execute(
            """
            SELECT u.*
            FROM session_tokens s
            JOIN users u ON u.id = s.user_id
            WHERE s.token_hash = ?
              AND s.expires_at > ?
              AND u.is_active = 1
            """,
            (token_hash, iso_now()),
        ).fetchone()
        if row is None:
            raise HTTPException(status_code=401, detail="Invalid session")
        return row
    finally:
        conn.close()


def fetch_cotizacion(conn: sqlite3.Connection, cotizacion_id: int) -> sqlite3.Row:
    row = conn.execute("SELECT * FROM cotizaciones WHERE id = ?", (cotizacion_id,)).fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail="Cotizacion not found")
    return row


def fetch_cotizacion_items(conn: sqlite3.Connection, cotizacion_id: int) -> list[sqlite3.Row]:
    return conn.execute(
        "SELECT * FROM cotizacion_items WHERE cotizacion_id = ? ORDER BY position ASC",
        (cotizacion_id,),
    ).fetchall()


def build_excel_for_quote(cotizacion: sqlite3.Row, items: list[sqlite3.Row], output_path: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = TARGET_SHEET_NAME

    thin = Side(border_style="thin", color="111111")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    dark_fill = PatternFill("solid", fgColor="2F2F2F")
    light_fill = PatternFill("solid", fgColor="E8E8E8")
    white_bold = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)

    column_widths = {
        "B": 25,
        "C": 10,
        "D": 14,
        "E": 47,
        "F": 8,
        "G": 10,
        "H": 11,
        "I": 10,
        "J": 11,
        "K": 11,
        "L": 11,
    }
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    # Heading
    sheet.merge_cells("B2:C2")
    sheet["B2"] = "Version: 03"
    sheet.merge_cells("B3:C3")
    sheet["B3"] = "Pagina: 1"

    sheet.merge_cells("I2:M2")
    sheet["I2"] = "APPCOT"
    sheet["I2"].font = Font(bold=True, size=18)
    sheet.merge_cells("I3:M3")
    sheet["I3"] = "Codigo: AM-400-000"

    sheet.merge_cells("B5:D5")
    sheet["B5"] = cotizacion["company_name"]
    sheet["B5"].fill = dark_fill
    sheet["B5"].font = white_bold
    sheet["B5"].alignment = left
    sheet["B5"].border = border

    attention_name = cotizacion["full_name"] or "Nombre y Apellido"
    sheet.merge_cells("E5:H5")
    sheet["E5"] = f"Attn: {attention_name}"
    sheet["E5"].fill = dark_fill
    sheet["E5"].font = white_bold
    sheet["E5"].alignment = center
    sheet["E5"].border = border

    sheet.merge_cells("I5:J5")
    sheet["I5"] = "Fecha:"
    sheet["I5"].fill = dark_fill
    sheet["I5"].font = white_bold
    sheet["I5"].alignment = center
    sheet["I5"].border = border

    sheet.merge_cells("K5:M5")
    sheet["K5"] = get_today_date_spanish()
    sheet["K5"].fill = dark_fill
    sheet["K5"].font = Font(color="F28C28", bold=True)
    sheet["K5"].alignment = center
    sheet["K5"].border = border

    header_row = 7
    sheet.row_dimensions[header_row].height = 42
    headers = {
        "B": "Estructura",
        "C": "Tapa /\nFondo",
        "D": "Linea/Producto",
        "E": "Descripcion del Material",
        "F": "Ancho\n(mm)",
        "G": "Longitud\nde bobina\n(m)",
        "H": "Volumen\nanual\nproyectado\n(mts)",
        "I": "Escala\ncotizada\n(mts)",
        "J": "Precio metro\n(USD)",
        "K": "Precio bobina\n(USD)",
        "L": "Precio Km\n(USD)",
    }
    for col, label in headers.items():
        cell = sheet[f"{col}{header_row}"]
        cell.value = label
        cell.font = bold
        cell.fill = light_fill
        cell.alignment = center
        cell.border = border

    monthly_value = parse_excel_number(cotizacion["monthly_meters"])
    product_label = cotizacion["product_name"] or "Flex GL"
    commission = float(cotizacion["commission_factor"] or DEFAULT_COMMISSION_FACTOR)
    line_product = cotizacion["line_product"] or ""

    data_start_row = 8
    for index, item in enumerate(items[:MAX_ITEMS]):
        row = data_start_row + index
        item_type = (item["type"] or "TAPA").strip().upper()
        item_type = item_type if item_type in {"TAPA", "FONDO"} else "TAPA"
        calibre = str(item["calibre"]).strip()
        width_value = parse_excel_number(item["width"])
        barrier = (item["barrier_type"] or "alta").strip().lower()
        seal = (item["seal_type"] or "hermetico").strip().lower()
        barrier_text = "mediana barrera" if barrier == "mediana" else "alta barrera"
        seal_text = "pelable" if seal == "pelable" else "hermético"
        milesimas = get_milesimas_for_material(product_label, calibre)
        material_price = (
            float(item["price_override_p100"])
            if item["price_override_p100"] is not None
            else get_price_for_material(product_label, calibre)
        )

        sheet[f"B{row}"] = f"{product_label} {calibre}".strip()
        sheet[f"C{row}"] = item_type
        sheet[f"D{row}"] = line_product
        if milesimas:
            sheet[f"E{row}"] = (
                f"Material coextruido y laminado, {barrier_text}, sello {seal_text} {milesimas} mil de espesor"
            )
        else:
            sheet[f"E{row}"] = f"Material coextruido y laminado, {barrier_text}, sello {seal_text}"
        sheet[f"F{row}"] = width_value if width_value is not None else ""
        sheet[f"G{row}"] = 914
        sheet[f"H{row}"] = "TBD"
        sheet[f"I{row}"] = monthly_value if monthly_value is not None else ""

        if material_price is not None and width_value is not None:
            qmil = 100000 / float(width_value)
            if qmil > 0:
                pmil = material_price / qmil
                pbase = pmil * commission
                price_km = pbase * 1000
                price_m = round(price_km / 1000, 3)
                price_bobina = round(price_m * 914, 2)
                sheet[f"J{row}"] = price_m
                sheet[f"K{row}"] = price_bobina
                sheet[f"L{row}"] = round(price_km, 2)
            else:
                sheet[f"J{row}"] = ""
                sheet[f"K{row}"] = ""
                sheet[f"L{row}"] = ""
        else:
            sheet[f"J{row}"] = ""
            sheet[f"K{row}"] = ""
            sheet[f"L{row}"] = ""

        for col in headers.keys():
            cell = sheet[f"{col}{row}"]
            cell.border = border
            cell.alignment = left if col in {"B", "D", "E"} else center
        sheet[f"I{row}"].number_format = "#,##0"
        sheet[f"J{row}"].number_format = "$#,##0.000"
        sheet[f"K{row}"].number_format = "$#,##0.00"
        sheet[f"L{row}"].number_format = "$#,##0.00"

    used_count = len(items[:MAX_ITEMS])
    for index in range(used_count, MAX_ITEMS):
        clear_product_row(sheet, data_start_row + index, "B", "L")

    footer_start = data_start_row + max(used_count, 1) + 2
    footer_end = footer_start + 4
    sheet.merge_cells(f"B{footer_start}:I{footer_start + 1}")
    sheet[f"B{footer_start}"] = (
        "Los precios anteriores son en USD(*), al tipo de cambio del dia de la facturacion, "
        "no incluye IVA y son DDP, credito: Por definir."
    )
    sheet[f"B{footer_start}"].alignment = left
    sheet[f"B{footer_start}"].border = border

    sheet.merge_cells(f"B{footer_start + 2}:I{footer_end}")
    sheet[f"B{footer_start + 2}"] = (
        "Consulte los terminos y condiciones en la siguiente liga:\n"
        "Terms and Conditions | APPCOT\n\n"
        "Debido a la situacion actual de las materias primas y a las considerables "
        "fluctuaciones de las mismas, nos reservamos el derecho de ajustar los precios.\n"
        "La presente cancela cualquier cotizacion anterior y los precios son vigentes "
        "durante un periodo de 15 dias."
    )
    sheet[f"B{footer_start + 2}"].alignment = left
    sheet[f"B{footer_start + 2}"].border = border

    sheet.merge_cells(f"J{footer_start}:L{footer_end}")
    sheet[f"J{footer_start}"] = (
        "APPCOT\n"
        "Aldo Manzur Coronel\n"
        "aldo.manzur@mx.multivac.com\n"
        "Celular 55 3232 7977\n\n"
        "[QR Placeholder]"
    )
    sheet[f"J{footer_start}"].alignment = center
    sheet[f"J{footer_start}"].font = bold
    sheet[f"J{footer_start}"].border = border

    workbook.save(output_path)
    workbook.close()


def build_quote_pdf_bytes(cotizacion: sqlite3.Row, items: list[sqlite3.Row]) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )
    styles = getSampleStyleSheet()
    elements: list[Any] = []

    title = Paragraph(f"<b>Cotizacion #{cotizacion['id']}</b>", styles["Title"])
    subtitle = Paragraph(
        f"Empresa: <b>{cotizacion['company_name']}</b> &nbsp;&nbsp; | &nbsp;&nbsp; Atn: <b>{cotizacion['full_name']}</b>",
        styles["Normal"],
    )
    meta = Paragraph(
        f"Fecha: {get_today_date_spanish()} &nbsp;&nbsp; | &nbsp;&nbsp; Linea/Producto: {cotizacion['line_product']}",
        styles["Normal"],
    )
    elements.extend([title, Spacer(1, 0.15 * inch), subtitle, Spacer(1, 0.07 * inch), meta, Spacer(1, 0.2 * inch)])

    header = [
        "Item",
        "Tipo",
        "Calibre",
        "Ancho",
        "Barrera",
        "Sello",
        "Precio producto",
        "Precio metro",
        "Precio bobina",
        "Precio km",
    ]
    table_data: list[list[str]] = [header]

    product_label = cotizacion["product_name"] or "Flex GL"
    commission = float(cotizacion["commission_factor"] or DEFAULT_COMMISSION_FACTOR)
    for idx, item in enumerate(items[:MAX_ITEMS], start=1):
        calibre = str(item["calibre"]).strip()
        width_value = parse_excel_number(item["width"])
        barrier = (item["barrier_type"] or "alta").strip().lower()
        seal = (item["seal_type"] or "hermetico").strip().lower()
        barrier_text = "Mediana" if barrier == "mediana" else "Alta"
        seal_text = "Pelable" if seal == "pelable" else "Hermetico"
        material_price = (
            float(item["price_override_p100"])
            if item["price_override_p100"] is not None
            else get_price_for_material(product_label, calibre)
        )

        price_m = None
        price_bobina = None
        price_km = None
        if material_price is not None and width_value is not None:
            qmil = 100000 / float(width_value)
            if qmil > 0:
                pmil = material_price / qmil
                pbase = pmil * commission
                price_km = pbase * 1000
                price_m = round(price_km / 1000, 3)
                price_bobina = round(price_m * 914, 2)

        table_data.append(
            [
                str(idx),
                (item["type"] or "").upper(),
                calibre,
                str(width_value if width_value is not None else ""),
                barrier_text,
                seal_text,
                f"${material_price:,.2f}" if material_price is not None else "-",
                f"${price_m:,.3f}" if price_m is not None else "-",
                f"${price_bobina:,.2f}" if price_bobina is not None else "-",
                f"${price_km:,.2f}" if price_km is not None else "-",
            ]
        )

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F2F2F")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 0.22 * inch))
    elements.append(
        Paragraph(
            "Este documento fue generado para revision comercial. Los precios son expresados en USD.",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 0.15 * inch))
    elements.append(Paragraph("<b>Atentamente,</b>", styles["Normal"]))
    elements.append(Paragraph("<b>Aldo Manzur Coronel</b>", styles["Normal"]))
    elements.append(Paragraph("Asesor comercial | AppCot", styles["Normal"]))
    elements.append(Paragraph("aldo.manzur@mx.multivac.com", styles["Normal"]))
    elements.append(Paragraph("Cel. 55 3232 7977", styles["Normal"]))

    doc.build(elements)
    return buffer.getvalue()


def build_quote_email_html(cotizacion_id: int, company_name: str, full_name: str) -> str:
    return f"""
    <div style="font-family: Arial, Helvetica, sans-serif; background: #f4f6f8; padding: 24px;">
      <div style="max-width: 680px; margin: 0 auto; background: white; border-radius: 12px; overflow: hidden; border: 1px solid #e9ecef;">
        <div style="background: #111827; color: #fff; padding: 18px 24px;">
          <h2 style="margin: 0; font-size: 20px;">Cotizacion aprobada</h2>
        </div>
        <div style="padding: 24px;">
          <p style="margin-top: 0; font-size: 16px; color: #111827;">Hola,</p>
          <p style="font-size: 15px; color: #374151;">
            Tu cotizacion <strong>#{cotizacion_id}</strong> fue aprobada.
          </p>
          <p style="font-size: 15px; color: #374151; line-height: 1.6;">
            <strong>Empresa:</strong> {company_name}<br/>
            <strong>Atn:</strong> {full_name}
          </p>
          <p style="font-size: 15px; color: #374151;">
            Adjuntamos el PDF de la cotizacion para su revision.
          </p>

          <div style="margin-top: 28px; padding-top: 16px; border-top: 1px solid #e5e7eb;">
            <p style="margin: 0; color: #111827; font-weight: bold;">Aldo Manzur Coronel</p>
            <p style="margin: 6px 0 0 0; color: #6b7280;">Asesor comercial | AppCot</p>
            <p style="margin: 4px 0 0 0; color: #6b7280;">aldo.manzur@mx.multivac.com</p>
            <p style="margin: 4px 0 0 0; color: #6b7280;">Cel. 55 3232 7977</p>
          </div>
        </div>
      </div>
    </div>
    """


@app.post("/api/cotizaciones")
def create_cotizacion(payload: CotizacionCreateRequest) -> dict:
    logger.info("create_cotizacion | company=%s | emails=%s | items=%d", payload.company_name, payload.emails, len(payload.items))
    if not payload.items:
        raise HTTPException(status_code=400, detail="At least one item is required")
    if len(payload.items) > MAX_ITEMS:
        raise HTTPException(status_code=400, detail=f"Maximum {MAX_ITEMS} items allowed")

    cleaned_emails = [str(email).strip().lower() for email in payload.emails if str(email).strip()]
    unique_emails: list[str] = []
    for email in cleaned_emails:
        if "@" not in email:
            raise HTTPException(status_code=400, detail=f'Correo invalido: "{email}"')
        if email not in unique_emails:
            unique_emails.append(email)
    if not unique_emails:
        raise HTTPException(status_code=400, detail="Debes agregar al menos un correo para enviar la cotizacion")

    conn = db_connect()
    now = iso_now()
    try:
        cur = conn.execute(
            """
            INSERT INTO cotizaciones (
              status, full_name, company_name, emails_json, line_product, monthly_meters,
              product_name, commission_factor, review_notes, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "pending",
                payload.full_name.strip(),
                payload.company_name.strip(),
                json.dumps(unique_emails, ensure_ascii=False),
                payload.line_product.strip(),
                payload.monthly_meters,
                payload.product_name.strip() or "Flex GL",
                DEFAULT_COMMISSION_FACTOR,
                None,
                now,
                now,
            ),
        )
        cotizacion_id = cur.lastrowid

        for position, item in enumerate(payload.items, start=1):
            width_value = parse_excel_number(item.width)
            if width_value is None:
                raise HTTPException(status_code=400, detail=f"Item #{position} width is invalid")
            conn.execute(
                """
                INSERT INTO cotizacion_items (
                  cotizacion_id, position, type, calibre, width, barrier_type, seal_type, price_override_p100
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    cotizacion_id,
                    position,
                    item.type.strip().upper(),
                    item.calibre.strip(),
                    float(width_value),
                    item.barrierType.strip().lower(),
                    item.sealType.strip().lower(),
                    item.priceOverrideP100,
                ),
            )
        conn.commit()
    finally:
        conn.close()

    logger.info("Cotizacion created | id=%s | company=%s", cotizacion_id, payload.company_name)
    return {
        "id": cotizacion_id,
        "status": "pending",
        "message": "Tu solicitud fue enviada. Un miembro del personal terminara la cotizacion y te llegara por correo.",
    }


@app.post("/api/auth/magic-link/request")
def request_magic_link(payload: MagicLinkRequest) -> dict:
    email = payload.email.strip().lower()
    logger.info("magic-link/request | email=%s", email)
    if "@" not in email:
        logger.warning("magic-link/request rejected | invalid email=%s", email)
        raise HTTPException(status_code=400, detail="Invalid email")

    conn = db_connect()
    now = iso_now()
    try:
        user = conn.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
        if user is None:
            cur = conn.execute(
                "INSERT INTO users (email, name, role, is_active, created_at) VALUES (?, ?, ?, ?, ?)",
                (email, payload.name or email.split("@")[0], "cotizador", 1, now),
            )
            user_id = cur.lastrowid
        else:
            user_id = user["id"]
            if payload.name:
                conn.execute("UPDATE users SET name = ? WHERE id = ?", (payload.name, user_id))

        raw_token = secrets.token_urlsafe(32)
        token_hash = hash_token(raw_token)
        expires_at = (utcnow() + timedelta(minutes=15)).isoformat()
        conn.execute(
            """
            INSERT INTO magic_link_tokens (user_id, token_hash, expires_at, used_at, created_at)
            VALUES (?, ?, ?, NULL, ?)
            """,
            (user_id, token_hash, expires_at, now),
        )
        conn.commit()
    finally:
        conn.close()

    logger.info("magic-link/request OK | email=%s | expires_at=%s", email, expires_at)
    return {
        "ok": True,
        "message": "Magic link generado (stub).",
        "magic_link_token": raw_token,
        "expires_at": expires_at,
    }


@app.post("/api/auth/magic-link/verify")
def verify_magic_link(payload: MagicLinkVerifyRequest, response: Response) -> dict:
    logger.info("magic-link/verify attempt")
    token_hash = hash_token(payload.token.strip())
    now = iso_now()
    conn = db_connect()
    try:
        token_row = conn.execute(
            """
            SELECT * FROM magic_link_tokens
            WHERE token_hash = ? AND used_at IS NULL AND expires_at > ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (token_hash, now),
        ).fetchone()
        if token_row is None:
            logger.warning("magic-link/verify failed | token invalid or expired")
            raise HTTPException(status_code=401, detail="Invalid or expired token")

        conn.execute("UPDATE magic_link_tokens SET used_at = ? WHERE id = ?", (now, token_row["id"]))
        session_raw = secrets.token_urlsafe(32)
        session_hash = hash_token(session_raw)
        session_exp = (utcnow() + timedelta(days=7)).isoformat()
        conn.execute(
            "INSERT INTO session_tokens (user_id, token_hash, expires_at, created_at) VALUES (?, ?, ?, ?)",
            (token_row["user_id"], session_hash, session_exp, now),
        )
        conn.commit()

        user = conn.execute("SELECT * FROM users WHERE id = ?", (token_row["user_id"],)).fetchone()
        logger.info("magic-link/verify OK | user_id=%s | email=%s", user["id"], user["email"])
    finally:
        conn.close()

    response.set_cookie(
        key=SESSION_COOKIE_NAME,
        value=session_raw,
        httponly=True,
        samesite="lax",
        secure=False,
        max_age=7 * 24 * 60 * 60,
        path="/",
    )
    return {
        "ok": True,
        "user": {
            "id": user["id"],
            "email": user["email"],
            "name": user["name"],
            "role": user["role"],
        },
    }


@app.post("/api/auth/logout")
def logout(response: Response, session_cookie: str | None = Cookie(default=None, alias=SESSION_COOKIE_NAME)) -> dict:
    if session_cookie:
        conn = db_connect()
        try:
            conn.execute("DELETE FROM session_tokens WHERE token_hash = ?", (hash_token(session_cookie),))
            conn.commit()
        finally:
            conn.close()
    response.delete_cookie(SESSION_COOKIE_NAME, path="/")
    return {"ok": True}


@app.get("/api/auth/me")
def auth_me(session_cookie: str | None = Cookie(default=None, alias=SESSION_COOKIE_NAME)) -> dict:
    user = get_current_user(session_cookie)
    return {
        "id": user["id"],
        "email": user["email"],
        "name": user["name"],
        "role": user["role"],
    }


@app.get("/api/cotizaciones")
def list_cotizaciones(
    status: str | None = Query(default=None),
    session_cookie: str | None = Cookie(default=None, alias=SESSION_COOKIE_NAME),
) -> dict:
    _ = get_current_user(session_cookie)
    conn = db_connect()
    try:
        if status == "completed":
            rows = conn.execute(
                """
                SELECT id, status, full_name, company_name, line_product, created_at, updated_at
                FROM cotizaciones
                WHERE status IN ('completed', 'approved')
                ORDER BY id DESC
                """
            ).fetchall()
        elif status:
            rows = conn.execute(
                """
                SELECT id, status, full_name, company_name, line_product, created_at, updated_at
                FROM cotizaciones
                WHERE status = ?
                ORDER BY id DESC
                """,
                (status,),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT id, status, full_name, company_name, line_product, created_at, updated_at
                FROM cotizaciones
                ORDER BY id DESC
                """
            ).fetchall()
        return {"items": [dict(r) for r in rows]}
    finally:
        conn.close()


@app.get("/api/cotizaciones/{cotizacion_id}")
def get_cotizacion(cotizacion_id: int, session_cookie: str | None = Cookie(default=None, alias=SESSION_COOKIE_NAME)) -> dict:
    _ = get_current_user(session_cookie)
    conn = db_connect()
    try:
        cotizacion = fetch_cotizacion(conn, cotizacion_id)
        items = fetch_cotizacion_items(conn, cotizacion_id)
        result = dict(cotizacion)
        result["emails"] = json.loads(result["emails_json"])
        product_name = result.get("product_name") or "Flex GL"
        enriched_items: list[dict[str, Any]] = []
        for row in items:
            item = dict(row)
            base_price = get_price_for_material(product_name, str(item.get("calibre", "")).strip())
            override_price = item.get("price_override_p100")
            effective_price = float(override_price) if override_price is not None else base_price
            item["base_price_p100"] = base_price
            item["effective_price_p100"] = effective_price
            enriched_items.append(item)
        result["items"] = enriched_items
        return result
    finally:
        conn.close()


@app.patch("/api/cotizaciones/{cotizacion_id}")
def update_cotizacion(
    cotizacion_id: int,
    payload: CotizacionUpdateRequest,
    session_cookie: str | None = Cookie(default=None, alias=SESSION_COOKIE_NAME),
) -> dict:
    user = get_current_user(session_cookie)
    logger.info("update_cotizacion | id=%s | by=%s | fields=%s", cotizacion_id, user["email"],
        [k for k, v in payload.model_dump().items() if v is not None])
    conn = db_connect()
    now = iso_now()
    try:
        _ = fetch_cotizacion(conn, cotizacion_id)
        updates: list[str] = []
        values: list[Any] = []
        if payload.commissionFactor is not None:
            updates.append("commission_factor = ?")
            values.append(payload.commissionFactor)
        if payload.reviewNotes is not None:
            updates.append("review_notes = ?")
            values.append(payload.reviewNotes)
        if payload.lineProduct is not None:
            updates.append("line_product = ?")
            values.append(payload.lineProduct)
        if payload.monthlyMeters is not None:
            updates.append("monthly_meters = ?")
            values.append(payload.monthlyMeters)
        if payload.emails is not None:
            updates.append("emails_json = ?")
            values.append(json.dumps(payload.emails))

        if updates:
            updates.append("updated_at = ?")
            values.append(now)
            values.append(cotizacion_id)
            conn.execute(f"UPDATE cotizaciones SET {', '.join(updates)} WHERE id = ?", values)

        if payload.items:
            for item_patch in payload.items:
                item_id = item_patch.get("id")
                if not item_id:
                    continue
                item_updates: list[str] = []
                item_values: list[Any] = []
                if "width" in item_patch:
                    width_val = parse_excel_number(item_patch["width"])
                    if width_val is None:
                        raise HTTPException(status_code=400, detail=f"Invalid width for item {item_id}")
                    item_updates.append("width = ?")
                    item_values.append(float(width_val))
                if "calibre" in item_patch and item_patch["calibre"]:
                    item_updates.append("calibre = ?")
                    item_values.append(str(item_patch["calibre"]).strip())
                if "barrier_type" in item_patch:
                    item_updates.append("barrier_type = ?")
                    item_values.append(str(item_patch["barrier_type"]).strip().lower())
                if "seal_type" in item_patch:
                    item_updates.append("seal_type = ?")
                    item_values.append(str(item_patch["seal_type"]).strip().lower())
                if "price_override_p100" in item_patch:
                    override = item_patch["price_override_p100"]
                    item_updates.append("price_override_p100 = ?")
                    item_values.append(float(override) if override not in (None, "") else None)
                if item_updates:
                    item_values.extend([cotizacion_id, item_id])
                    conn.execute(
                        f"UPDATE cotizacion_items SET {', '.join(item_updates)} WHERE cotizacion_id = ? AND id = ?",
                        item_values,
                    )
            conn.execute("UPDATE cotizaciones SET updated_at = ? WHERE id = ?", (now, cotizacion_id))

        conn.commit()
    finally:
        conn.close()

    return {"ok": True}


@app.delete("/api/cotizaciones/{cotizacion_id}")
def delete_cotizacion(
    cotizacion_id: int, session_cookie: str | None = Cookie(default=None, alias=SESSION_COOKIE_NAME)
) -> dict:
    _ = get_current_user(session_cookie)
    conn = db_connect()
    try:
        _ = fetch_cotizacion(conn, cotizacion_id)
        conn.execute("DELETE FROM cotizaciones WHERE id = ?", (cotizacion_id,))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True, "deleted_id": cotizacion_id}


@app.get("/api/cotizaciones/{cotizacion_id}/excel")
def preview_cotizacion_excel(
    cotizacion_id: int, session_cookie: str | None = Cookie(default=None, alias=SESSION_COOKIE_NAME)
) -> FileResponse:
    _ = get_current_user(session_cookie)
    conn = db_connect()
    try:
        cotizacion = fetch_cotizacion(conn, cotizacion_id)
        items = fetch_cotizacion_items(conn, cotizacion_id)
    finally:
        conn.close()

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        output_path = temp_file.name
    try:
        build_excel_for_quote(cotizacion, items, output_path)
    except Exception as exc:  # noqa: BLE001
        remove_temp_file(output_path)
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel: {exc}") from exc

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"cotizacion-{cotizacion_id}.xlsx",
        background=BackgroundTask(remove_temp_file, output_path),
    )


@app.post("/api/cotizaciones/{cotizacion_id}/approve")
def approve_cotizacion(
    cotizacion_id: int, session_cookie: str | None = Cookie(default=None, alias=SESSION_COOKIE_NAME)
) -> dict:
    user = get_current_user(session_cookie)
    logger.info("approve_cotizacion | id=%s | by=%s", cotizacion_id, user["email"])
    conn = db_connect()
    now = iso_now()
    try:
        cotizacion = fetch_cotizacion(conn, cotizacion_id)
        if cotizacion["status"] != "completed":
            conn.execute(
                """
                UPDATE cotizaciones
                SET status = 'completed', approved_at = ?, approved_by_user_id = ?, updated_at = ?
                WHERE id = ?
                """,
                (now, user["id"], now, cotizacion_id),
            )

        recipients = [str(email).strip() for email in json.loads(cotizacion["emails_json"]) if str(email).strip()]
        if not recipients and EMAIL_REPLY_TO:
            logger.info("approve_cotizacion | id=%s | no recipients in DB, falling back to EMAIL_REPLY_TO=%s", cotizacion_id, EMAIL_REPLY_TO)
            recipients = [EMAIL_REPLY_TO]
        if not recipients:
            logger.error("approve_cotizacion | id=%s | no recipients configured", cotizacion_id)
            raise HTTPException(status_code=400, detail="No hay correos configurados para esta cotizacion")
        logger.info("approve_cotizacion | id=%s | recipients=%s | email_mode=%s", cotizacion_id, recipients, EMAIL_MODE)

        subject = f"Cotizacion #{cotizacion_id} aprobada"
        body_preview = (
            f"Hola,\n\nLa cotizacion #{cotizacion_id} fue aprobada.\n"
            f"Empresa: {cotizacion['company_name']}\n"
            f"Atn: {cotizacion['full_name']}\n\n"
            "Adjuntamos el PDF de la cotizacion para su revision.\n\n"
            "Atentamente,\nAldo Manzur Coronel"
        )
        email_log_suffix = "stub"
        if EMAIL_MODE == "resend":
            try:
                items_for_pdf = fetch_cotizacion_items(conn, cotizacion_id)
                pdf_bytes = build_quote_pdf_bytes(cotizacion, items_for_pdf)
                attachments = [
                    {
                        "filename": f"cotizacion-{cotizacion_id}.pdf",
                        "content": base64.b64encode(pdf_bytes).decode("utf-8"),
                    }
                ]
                provider_ids: list[str] = []
                for recipient in recipients:
                    resend_result = send_email_with_resend(
                        recipients=[recipient],
                        subject=subject,
                        html_body=build_quote_email_html(
                            cotizacion_id=cotizacion_id,
                            company_name=cotizacion["company_name"],
                            full_name=cotizacion["full_name"],
                        ),
                        text_body=body_preview,
                        attachments=attachments,
                    )
                    provider_id = resend_result.get("provider_id")
                    if provider_id:
                        provider_ids.append(str(provider_id))
                email_log_suffix = f"resend_sent ids={','.join(provider_ids)}" if provider_ids else "resend_sent"
            except RuntimeError as exc:
                logger.error("approve_cotizacion | id=%s | email send failed: %s", cotizacion_id, exc)
                raise HTTPException(status_code=502, detail=f"No se pudo enviar correo con Resend: {exc}") from exc

        conn.execute(
            """
            INSERT INTO email_stub_logs (cotizacion_id, recipients_json, subject, body_preview, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                cotizacion_id,
                json.dumps(recipients, ensure_ascii=False),
                subject,
                f"{body_preview}\n\n[delivery={email_log_suffix}]",
                now,
            ),
        )
        conn.commit()
        logger.info("approve_cotizacion | id=%s | done | delivery=%s", cotizacion_id, email_log_suffix)
    finally:
        conn.close()
    return {"ok": True, "status": "completed", "email_mode": EMAIL_MODE}


@app.get("/api/materiales/tapas/calibres")
def get_tapa_calibres(
    material_name: Annotated[str, Query(alias="materialName")] = "Flex GL",
) -> dict:
    material = get_material_record(material_name)
    if material is None:
        raise HTTPException(status_code=404, detail=f'Material "{material_name}" not found')

    prices_by_micras = material.get("prices_by_micras", {})
    calibres = []
    for micras_key, values in prices_by_micras.items():
        micras = int(micras_key)
        milesimas_raw = values.get("espesor_milesimas")
        milesimas = f"{float(milesimas_raw):.1f}"
        calibres.append({"micras": micras, "milesimas": milesimas})
    calibres.sort(key=lambda item: item["micras"])
    return {"material": material_name, "calibres": calibres}


@app.get("/api/debug/email-status/{email_id}")
def get_email_status(email_id: str) -> dict:
    """Query Resend API for the delivery status of a sent email."""
    if not RESEND_API_KEY:
        raise HTTPException(status_code=503, detail="RESEND_API_KEY not configured")
    url = f"https://api.resend.com/emails/{email_id}"
    request = urllib.request.Request(
        url,
        headers={
            "Authorization": f"Bearer {RESEND_API_KEY}",
            "User-Agent": "appcot-backend/0.1",
        },
        method="GET",
    )
    try:
        with urllib.request.urlopen(request, timeout=15) as response:  # noqa: S310
            parsed = json.loads(response.read().decode("utf-8"))
            logger.info("email_status | id=%s | status=%s | last_event=%s", email_id, parsed.get("last_event"), parsed.get("created_at"))
            return parsed
    except urllib.error.HTTPError as exc:
        details = exc.read().decode("utf-8", errors="replace")
        raise HTTPException(status_code=exc.code, detail=details) from exc


if __name__ == "__main__":
    uvicorn.run("main:app", host="localhost", port=8009, reload=True)
"""
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Annotated
from datetime import datetime
import json

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, ConfigDict, Field
from starlette.background import BackgroundTask
import uvicorn


BASE_DIR = Path(__file__).resolve().parent
PRICES_PATH = BASE_DIR / "materiales" / "prices.json"
TARGET_SHEET_NAME = "Formato No. 1"
MAX_ITEMS = 4
COMMISSION_FACTOR = 1.15
SPANISH_MONTHS = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}


def remove_temp_file(path: str) -> None:
    try:
        Path(path).unlink(missing_ok=True)
    except OSError:
        pass


def get_today_date_spanish() -> str:
    today = datetime.now()
    month_name = SPANISH_MONTHS[today.month]
    return f"{today.day} de {month_name} de {today.year}"


def parse_excel_number(value: str | int | float | None) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value

    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        parsed = float(text)
        return int(parsed) if parsed.is_integer() else parsed
    except ValueError:
        return None


def clear_product_row(sheet, row: int, start_col: str = "B", end_col: str = "M") -> None:
    for col_ord in range(ord(start_col), ord(end_col) + 1):
        sheet[f"{chr(col_ord)}{row}"] = None


def get_material_record(material_name: str) -> dict | None:
    if not PRICES_PATH.exists():
        return None
    try:
        with PRICES_PATH.open("r", encoding="utf-8") as file:
            prices_data = json.load(file)
    except Exception:  # noqa: BLE001
        return None

    tapas = prices_data.get("materiales", {}).get("tapas", [])
    return next(
        (item for item in tapas if str(item.get("name", "")).strip().lower() == material_name.strip().lower()),
        None,
    )


def get_milesimas_for_material(material_name: str, calibre_micras: str) -> str | None:
    material = get_material_record(material_name)
    if material is None:
        return None
    try:
        micras_key = str(int(float(calibre_micras.strip())))
    except ValueError:
        return None
    values = material.get("prices_by_micras", {}).get(micras_key)
    if values is None or values.get("espesor_milesimas") is None:
        return None
    return f"{float(values['espesor_milesimas']):.1f}"


def get_price_for_material(material_name: str, calibre_micras: str) -> float | None:
    material = get_material_record(material_name)
    if material is None:
        return None
    try:
        micras_key = str(int(float(calibre_micras.strip())))
    except ValueError:
        return None
    values = material.get("prices_by_micras", {}).get(micras_key)
    if values is None or values.get("price") is None:
        return None
    try:
        return float(values["price"])
    except (TypeError, ValueError):
        return None


class QuoteRequest(BaseModel):
    model_config = ConfigDict(populate_by_name=True)
    company_name: Annotated[str, Field(min_length=1, alias="companyName")]
    full_name: Annotated[str, Field(default="", alias="fullName")]
    product_name: Annotated[str, Field(default="AMILEN ML", alias="productName")]
    top_calibre: Annotated[str, Field(default="", alias="topCalibre")]
    product_side: Annotated[str, Field(default="TAPA", alias="productSide")]
    line_product: Annotated[str, Field(default="", alias="lineProduct")]
    barrier_type: Annotated[str, Field(default="alta", alias="barrierType")]
    seal_type: Annotated[str, Field(default="hermetico", alias="sealType")]
    item_type: Annotated[str, Field(default="", alias="itemType")]
    item_calibre: Annotated[str, Field(default="", alias="itemCalibre")]
    item_width: Annotated[str | int | float, Field(default="", alias="itemWidth")]
    monthly_scale: Annotated[str | int | float, Field(default="", alias="monthlyMeters")]
    items: list[dict] = Field(default_factory=list, alias="items", max_length=MAX_ITEMS)


app = FastAPI(title="Multivac Quote API", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:3001",
        "http://localhost:3002",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.post("/api/quotes/generate")
def generate_quote_excel(payload: QuoteRequest) -> FileResponse:
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        output_path = temp_file.name

    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = TARGET_SHEET_NAME

        thin = Side(border_style="thin", color="111111")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dark_fill = PatternFill("solid", fgColor="2F2F2F")
        light_fill = PatternFill("solid", fgColor="E8E8E8")
        white_bold = Font(color="FFFFFF", bold=True)
        bold = Font(bold=True)

        column_widths = {
            "B": 25,
            "C": 10,
            "D": 14,
            "E": 47,
            "F": 8,
            "G": 10,
            "H": 11,
            "I": 10,
            "J": 11,
            "K": 11,
            "L": 11,
        }
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        # Heading
        sheet.merge_cells("B2:C2")
        sheet["B2"] = "Version: 03"
        sheet.merge_cells("B3:C3")
        sheet["B3"] = "Pagina: 1"

        sheet.merge_cells("I2:M2")
        sheet["I2"] = "APPCOT"
        sheet["I2"].font = Font(bold=True, size=18)
        sheet.merge_cells("I3:M3")
        sheet["I3"] = "Codigo: AM-400-000"

        sheet.merge_cells("B5:D5")
        sheet["B5"] = payload.company_name
        sheet["B5"].fill = dark_fill
        sheet["B5"].font = white_bold
        sheet["B5"].alignment = left
        sheet["B5"].border = border

        attention_name = payload.full_name.strip() or "Nombre y Apellido"
        sheet.merge_cells("E5:H5")
        sheet["E5"] = f"Attn: {attention_name}"
        sheet["E5"].fill = dark_fill
        sheet["E5"].font = white_bold
        sheet["E5"].alignment = center
        sheet["E5"].border = border

        sheet.merge_cells("I5:J5")
        sheet["I5"] = "Fecha:"
        sheet["I5"].fill = dark_fill
        sheet["I5"].font = white_bold
        sheet["I5"].alignment = center
        sheet["I5"].border = border

        sheet.merge_cells("K5:M5")
        sheet["K5"] = get_today_date_spanish()
        sheet["K5"].fill = dark_fill
        sheet["K5"].font = Font(color="F28C28", bold=True)
        sheet["K5"].alignment = center
        sheet["K5"].border = border

        # Table header
        header_row = 7
        sheet.row_dimensions[header_row].height = 42
        headers = {
            "B": "Estructura",
            "C": "Tapa /\nFondo",
            "D": "Linea/Producto",
            "E": "Descripcion del Material",
            "F": "Ancho\n(mm)",
            "G": "Longitud\nde bobina\n(m)",
            "H": "Volumen\nanual\nproyectado\n(mts)",
            "I": "Escala\ncotizada\n(mts)",
            "J": "Precio metro\n(USD)",
            "K": "Precio bobina\n(USD)",
            "L": "Precio Km\n(USD)",
        }
        for col, label in headers.items():
            cell = sheet[f"{col}{header_row}"]
            cell.value = label
            cell.font = bold
            cell.fill = light_fill
            cell.alignment = center
            cell.border = border

        monthly_value = parse_excel_number(payload.monthly_scale)
        product_label = payload.product_name.strip() or "AMILEN ML"

        normalized_items = payload.items[:MAX_ITEMS]
        if not normalized_items:
            normalized_items = [
                {
                    "type": payload.item_type.strip() or payload.product_side.strip(),
                    "calibre": payload.item_calibre.strip() or payload.top_calibre.strip(),
                    "width": payload.item_width,
                    "barrierType": payload.barrier_type,
                    "sealType": payload.seal_type,
                }
            ]

        if not normalized_items:
            raise HTTPException(status_code=400, detail="At least one material item is required")

        for index, item in enumerate(normalized_items, start=1):
            calibre_value = str(item.get("calibre", "")).strip()
            width_value = parse_excel_number(item.get("width"))
            if not calibre_value or width_value is None:
                raise HTTPException(
                    status_code=400,
                    detail=f"Item #{index} must include both width and calibre",
                )

        data_start_row = 8
        for index, item in enumerate(normalized_items[:MAX_ITEMS]):
            row = data_start_row + index
            item_type = str(item.get("type", "")).strip().upper()
            item_type = item_type if item_type in {"TAPA", "FONDO"} else "TAPA"
            calibre = str(item.get("calibre", "")).strip()
            width_value = parse_excel_number(item.get("width"))
            barrier = str(item.get("barrierType", "alta")).strip().lower()
            seal = str(item.get("sealType", "hermetico")).strip().lower()
            barrier_text = "mediana barrera" if barrier == "mediana" else "alta barrera"
            seal_text = "pelable" if seal == "pelable" else "hermético"
            milesimas = get_milesimas_for_material(product_label, calibre)
            material_price = get_price_for_material(product_label, calibre)

            sheet[f"B{row}"] = f"{product_label} {calibre}".strip()
            sheet[f"C{row}"] = item_type
            sheet[f"D{row}"] = payload.line_product.strip()
            if milesimas:
                sheet[f"E{row}"] = (
                    f"Material coextruido y laminado, {barrier_text}, "
                    f"sello {seal_text} {milesimas} mil de espesor"
                )
            else:
                sheet[f"E{row}"] = f"Material coextruido y laminado, {barrier_text}, sello {seal_text}"

            sheet[f"F{row}"] = width_value if width_value is not None else ""
            sheet[f"G{row}"] = 914
            sheet[f"H{row}"] = "TBD"
            sheet[f"I{row}"] = monthly_value if monthly_value is not None else ""

            qmil = None
            pmil = None
            pbase = None
            price_km = None
            price_m = None
            price_bobina = None
            if material_price is not None and width_value is not None:
                qmil = 100000 / float(width_value)
                if qmil > 0:
                    pmil = material_price / qmil
                    pbase = pmil * COMMISSION_FACTOR
                    price_km = pbase * 1000
                    price_m = round(price_km / 1000, 3)
                    price_bobina = round(price_m * 914, 2)
                    sheet[f"J{row}"] = price_m
                    sheet[f"K{row}"] = price_bobina
                    sheet[f"L{row}"] = round(price_km, 2)
                else:
                    sheet[f"J{row}"] = ""
                    sheet[f"K{row}"] = ""
                    sheet[f"L{row}"] = ""
            else:
                sheet[f"J{row}"] = ""
                sheet[f"K{row}"] = ""
                sheet[f"L{row}"] = ""

            for col in headers.keys():
                cell = sheet[f"{col}{row}"]
                cell.border = border
                cell.alignment = left if col in {"B", "D", "E"} else center

            sheet[f"I{row}"].number_format = "#,##0"
            sheet[f"J{row}"].number_format = "$#,##0.000"
            sheet[f"K{row}"].number_format = "$#,##0.00"
            sheet[f"L{row}"].number_format = "$#,##0.00"

        used_count = len(normalized_items[:MAX_ITEMS])
        for index in range(used_count, MAX_ITEMS):
            clear_product_row(sheet, data_start_row + index, "B", "L")

        # Footer
        footer_start = data_start_row + max(used_count, 1) + 2
        footer_end = footer_start + 4
        sheet.merge_cells(f"B{footer_start}:I{footer_start + 1}")
        sheet[f"B{footer_start}"] = (
            "Los precios anteriores son en USD(*), al tipo de cambio del dia de la facturacion, "
            "no incluye IVA y son DDP, credito: Por definir."
        )
        sheet[f"B{footer_start}"].alignment = left
        sheet[f"B{footer_start}"].border = border

        sheet.merge_cells(f"B{footer_start + 2}:I{footer_end}")
        sheet[f"B{footer_start + 2}"] = (
            "Consulte los terminos y condiciones en la siguiente liga:\n"
            "Terms and Conditions | APPCOT\n\n"
            "Debido a la situacion actual de las materias primas y a las considerables "
            "fluctuaciones de las mismas, nos reservamos el derecho de ajustar los precios.\n"
            "La presente cancela cualquier cotizacion anterior y los precios son vigentes "
            "durante un periodo de 15 dias."
        )
        sheet[f"B{footer_start + 2}"].alignment = left
        sheet[f"B{footer_start + 2}"].border = border

        sheet.merge_cells(f"J{footer_start}:L{footer_end}")
        sheet[f"J{footer_start}"] = (
            "APPCOT\n"
            "Aldo Manzur Coronel\n"
            "aldo.manzur@mx.multivac.com\n"
            "Celular 55 3232 7977\n\n"
            "[QR Placeholder]"
        )
        sheet[f"J{footer_start}"].alignment = center
        sheet[f"J{footer_start}"].font = bold
        sheet[f"J{footer_start}"].border = border

        workbook.save(output_path)
        workbook.close()
    except HTTPException:
        remove_temp_file(output_path)
        raise
    except Exception as exc:  # noqa: BLE001
        remove_temp_file(output_path)
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel: {exc}") from exc

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="cotizacion-generada.xlsx",
        background=BackgroundTask(remove_temp_file, output_path),
    )


@app.get("/api/materiales/tapas/calibres")
def get_tapa_calibres(
    material_name: Annotated[str, Query(alias="materialName")] = "Amilen ML",
) -> dict:
    material = get_material_record(material_name)
    if material is None:
        raise HTTPException(status_code=404, detail=f'Material "{material_name}" not found')

    prices_by_micras = material.get("prices_by_micras", {})
    calibres = []
    for micras_key, values in prices_by_micras.items():
        micras = int(micras_key)
        milesimas_raw = values.get("espesor_milesimas")
        milesimas = f"{float(milesimas_raw):.1f}"
        calibres.append({"micras": micras, "milesimas": milesimas})

    calibres.sort(key=lambda item: item["micras"])
    return {"material": material_name, "calibres": calibres}


if __name__ == "__main__":
    uvicorn.run("main:app", host="127.0.0.1", port=8009, reload=True)
"""
